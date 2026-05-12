from __future__ import annotations

import re
import unicodedata
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

from voucher_generator.profiles import get_profile_config


def clean_text(value: Any) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    if text in {"", "-", "--", "N/A", "n/a"}:
        return None
    return text


def normalize_email(value: Any) -> Optional[str]:
    text = clean_text(value)
    return text.lower() if text else None


def normalize_phone(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, (int, float)) and int(value) == value:
        return str(int(value))
    return clean_text(value)


def normalize_int(value: Any) -> Optional[int]:
    if value in (None, ""):
        return None
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return None


def normalize_date(value: Any) -> Optional[str]:
    if value in (None, ""):
        return None

    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")

    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")

    if isinstance(value, (int, float)):
        try:
            serial = float(value)
            if 1 <= serial <= 60000:
                excel_epoch = datetime(1899, 12, 30)
                parsed = excel_epoch + timedelta(days=serial)
                return parsed.strftime("%Y-%m-%d")
        except Exception:
            pass

    text = str(value).strip()
    for fmt in (
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%m/%d/%Y",
        "%d-%m-%Y",
        "%d.%m.%Y",
        "%Y/%m/%d",
    ):
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue

    return text


def normalize_time(value: Any) -> Optional[str]:
    if value in (None, ""):
        return None

    if isinstance(value, datetime):
        return value.strftime("%H:%M")

    if isinstance(value, (int, float)):
        try:
            numeric_value = float(value)

            if 0 <= numeric_value < 1:
                total_minutes = round(numeric_value * 24 * 60)
                hours = total_minutes // 60
                minutes = total_minutes % 60
                return f"{hours:02d}:{minutes:02d}"

            if numeric_value.is_integer():
                return str(int(numeric_value))
        except Exception:
            pass

    text = clean_text(value)
    if not text:
        return None

    for fmt in ("%H:%M", "%H.%M", "%I:%M %p"):
        try:
            return datetime.strptime(text, fmt).strftime("%H:%M")
        except ValueError:
            continue

    try:
        parsed = datetime.strptime(text, "%H:%M:%S")
        return parsed.strftime("%H:%M")
    except ValueError:
        pass

    return text


def normalize_key_text(value: Any) -> Optional[str]:
    text = clean_text(value)
    if not text:
        return None
    return text.upper()


def build_passenger_key(
    passport_number: Optional[str],
    last_name: Optional[str],
    first_name: Optional[str],
    date_of_birth: Optional[str],
    excel_row_number: int,
) -> str:
    passport = normalize_key_text(passport_number)
    if passport:
        return passport

    natural_key = "|".join(
        [
            normalize_key_text(last_name) or "NO-LASTNAME",
            normalize_key_text(first_name) or "NO-FIRSTNAME",
            normalize_date(date_of_birth) or "NO-DOB",
        ]
    )
    return f"{natural_key}|ROW-{excel_row_number}"


def build_merged_lookup(ws) -> Dict[Tuple[int, int], Tuple[int, int, int, int]]:
    lookup: Dict[Tuple[int, int], Tuple[int, int, int, int]] = {}
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                lookup[(row, col)] = (min_row, min_col, max_row, max_col)
    return lookup


def get_effective_cell_value(ws, merged_lookup, row: int, col: Optional[int]) -> Any:
    if not col:
        return None

    merge_bounds = merged_lookup.get((row, col))
    if merge_bounds:
        min_row, min_col, _, _ = merge_bounds
        return ws.cell(min_row, min_col).value

    return ws.cell(row, col).value


def get_merge_anchor_id(merged_lookup, row: int, col: Optional[int]) -> Optional[str]:
    if not col:
        return None
    merge_bounds = merged_lookup.get((row, col))
    if not merge_bounds:
        return None
    min_row, min_col, max_row, max_col = merge_bounds
    return f"R{min_row}C{min_col}:R{max_row}C{max_col}"


def normalize_header(value: Any) -> Optional[str]:
    text = clean_text(value)
    if not text:
        return None
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.upper()
    text = re.sub(r"[^\w\s#]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text or None


def build_header_index(ws, merged_lookup, header_row: int) -> Dict[str, int]:
    header_index: Dict[str, int] = {}

    for col in range(1, ws.max_column + 1):
        raw_header = get_effective_cell_value(ws, merged_lookup, header_row, col)
        normalized = normalize_header(raw_header)
        if normalized and normalized not in header_index:
            header_index[normalized] = col

    return header_index


def resolve_columns(
    header_index: Dict[str, int],
    aliases: Dict[str, List[str]],
    required_fields: List[str],
    header_row: int,
) -> Dict[str, int]:
    resolved: Dict[str, int] = {}
    missing_required: List[str] = []

    for field_name, candidates in aliases.items():
        for candidate in candidates:
            normalized_candidate = normalize_header(candidate)
            if normalized_candidate and normalized_candidate in header_index:
                resolved[field_name] = header_index[normalized_candidate]
                break

        if field_name in required_fields and field_name not in resolved:
            missing_required.append(field_name)

    if missing_required:
        details = []
        for field_name in missing_required:
            aliases_text = ", ".join(aliases.get(field_name, []))
            details.append(f"- {field_name}: expected one of [{aliases_text}]")
        raise ValueError(
            "Missing required columns in header row "
            f"{header_row}:\n" + "\n".join(details)
        )

    return resolved


def get_field_value(
    ws,
    merged_lookup,
    row: int,
    resolved_columns: Dict[str, int],
    field_name: str,
) -> Any:
    col = resolved_columns.get(field_name)
    if not col:
        return None
    return get_effective_cell_value(ws, merged_lookup, row, col)


def get_flight_direction_from_group(value: Any) -> Optional[str]:
    normalized = normalize_header(value)
    if not normalized:
        return None

    if "FLIGHT" not in normalized:
        return None

    if "IDA" in normalized or "OUTBOUND" in normalized:
        return "outbound"

    if "VUELTA" in normalized or "REGRESO" in normalized or "RETURN" in normalized:
        return "return"

    return None


def get_flight_segment_number(value: Any) -> Optional[int]:
    normalized = normalize_header(value)
    if not normalized:
        return None

    match = re.search(r"FLIGHT\s*(\d+)", normalized)
    if not match:
        return None

    return normalize_int(match.group(1))


def get_flight_field_from_header(value: Any) -> Optional[str]:
    normalized = normalize_header(value)
    if not normalized:
        return None

    if "FLIGHT" in normalized and "NUMBER" in normalized:
        return "flight_number"

    if normalized in {"FROM TO", "FROMTO"} or normalized.startswith("FROM"):
        return "origin"

    if "DATE OF DEPARTURE" in normalized:
        return "departure_date"

    if "TIME OF DEPARTURE" in normalized:
        return "departure_time"

    if "TIME OF ARRIVAL" in normalized:
        return "arrival_time"

    if "DATE OF ARRIVAL" in normalized:
        return "arrival_date"

    if normalized == "AIRPORT":
        return "destination_airport"

    return None


def build_flight_column_map(
    ws,
    merged_lookup,
    group_row: int,
    header_row: int,
) -> Dict[str, Dict[int, Dict[str, int]]]:
    flight_columns: Dict[str, Dict[int, Dict[str, int]]] = {
        "outbound": {},
        "return": {},
    }

    for col in range(1, ws.max_column + 1):
        group_value = get_effective_cell_value(ws, merged_lookup, group_row, col)
        header_value = get_effective_cell_value(ws, merged_lookup, header_row, col)

        direction = get_flight_direction_from_group(group_value)
        segment_number = get_flight_segment_number(group_value)
        field_name = get_flight_field_from_header(header_value)

        if not direction or not segment_number or not field_name:
            continue

        flight_columns.setdefault(direction, {})
        flight_columns[direction].setdefault(segment_number, {})
        flight_columns[direction][segment_number][field_name] = col

    return flight_columns


def flight_segment_has_data(segment: Dict[str, Any]) -> bool:
    return any(
        segment.get(key)
        for key in (
            "flight_number",
            "origin",
            "destination_airport",
            "departure_date",
            "departure_time",
            "arrival_date",
            "arrival_time",
        )
    )


def extract_flight_segments_for_row(
    ws,
    merged_lookup,
    excel_row: int,
    flight_column_map: Dict[str, Dict[int, Dict[str, int]]],
) -> Dict[str, List[Dict[str, Any]]]:
    flights: Dict[str, List[Dict[str, Any]]] = {
        "outbound": [],
        "return": [],
    }

    for direction in ("outbound", "return"):
        raw_segments: List[Dict[str, Any]] = []

        for segment_number in sorted(flight_column_map.get(direction, {}).keys()):
            columns = flight_column_map[direction][segment_number]

            segment = {
                "source_segment_number": segment_number,
                "flight_number": clean_text(
                    get_effective_cell_value(
                        ws,
                        merged_lookup,
                        excel_row,
                        columns.get("flight_number"),
                    )
                ),
                "origin": clean_text(
                    get_effective_cell_value(
                        ws,
                        merged_lookup,
                        excel_row,
                        columns.get("origin"),
                    )
                ),
                "departure_date": normalize_date(
                    get_effective_cell_value(
                        ws,
                        merged_lookup,
                        excel_row,
                        columns.get("departure_date"),
                    )
                ),
                "departure_time": normalize_time(
                    get_effective_cell_value(
                        ws,
                        merged_lookup,
                        excel_row,
                        columns.get("departure_time"),
                    )
                ),
                "arrival_time": normalize_time(
                    get_effective_cell_value(
                        ws,
                        merged_lookup,
                        excel_row,
                        columns.get("arrival_time"),
                    )
                ),
                "arrival_date": normalize_date(
                    get_effective_cell_value(
                        ws,
                        merged_lookup,
                        excel_row,
                        columns.get("arrival_date"),
                    )
                ),
                "destination_airport": clean_text(
                    get_effective_cell_value(
                        ws,
                        merged_lookup,
                        excel_row,
                        columns.get("destination_airport"),
                    )
                ),
            }

            if flight_segment_has_data(segment):
                raw_segments.append(segment)

        for index, segment in enumerate(raw_segments, start=1):
            segment["segment_order"] = index
            flights[direction].append(segment)

    return flights


def read_effective_rows(
    xlsx_path: Path,
    sheet_name: Optional[str] = None,
    profile_name: str = "default",
) -> List[Dict[str, Any]]:
    profile = get_profile_config(profile_name)

    header_row = profile["header_row"]
    start_row = profile["start_row"]
    field_aliases = profile["field_aliases"]
    required_fields = profile["required_fields"]

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    merged_lookup = build_merged_lookup(ws)

    event_name = clean_text(ws.cell(row=1, column=1).value)

    header_index = build_header_index(ws, merged_lookup, header_row=header_row)
    resolved_columns = resolve_columns(
        header_index=header_index,
        aliases=field_aliases,
        required_fields=required_fields,
        header_row=header_row,
    )

    flight_column_map = build_flight_column_map(
        ws=ws,
        merged_lookup=merged_lookup,
        group_row=header_row - 1,
        header_row=header_row,
    )

    rows: List[Dict[str, Any]] = []

    for excel_row in range(start_row, ws.max_row + 1):
        raw_first_name = get_field_value(ws, merged_lookup, excel_row, resolved_columns, "first_name")
        raw_last_name = get_field_value(ws, merged_lookup, excel_row, resolved_columns, "last_name")
        raw_destination = get_field_value(ws, merged_lookup, excel_row, resolved_columns, "destination")
        raw_hotel_name = get_field_value(ws, merged_lookup, excel_row, resolved_columns, "hotel_name")
        raw_room = get_field_value(ws, merged_lookup, excel_row, resolved_columns, "room")
        raw_check_in = get_field_value(ws, merged_lookup, excel_row, resolved_columns, "check_in")
        raw_check_out = get_field_value(ws, merged_lookup, excel_row, resolved_columns, "check_out")
        raw_qty = get_field_value(ws, merged_lookup, excel_row, resolved_columns, "qty")

        first_name = clean_text(raw_first_name)
        last_name = clean_text(raw_last_name)
        destination = clean_text(raw_destination)
        hotel_name = clean_text(raw_hotel_name)
        room = clean_text(raw_room)
        check_in = normalize_date(raw_check_in)
        check_out = normalize_date(raw_check_out)
        qty = normalize_int(raw_qty)

        has_passenger = bool(first_name or last_name)
        has_voucher_context = any([destination, hotel_name, room, check_in, check_out, qty])

        if not has_passenger and not has_voucher_context:
            continue

        qty_merge_anchor = get_merge_anchor_id(
            merged_lookup,
            excel_row,
            resolved_columns.get("qty"),
        )
        row_number_merge_anchor = get_merge_anchor_id(
            merged_lookup,
            excel_row,
            resolved_columns.get("row_number"),
        )

        source_row_number = normalize_int(
            get_field_value(ws, merged_lookup, excel_row, resolved_columns, "row_number")
        )
        group_label = clean_text(
            get_field_value(ws, merged_lookup, excel_row, resolved_columns, "group_label")
        )
        date_of_birth = normalize_date(
            get_field_value(ws, merged_lookup, excel_row, resolved_columns, "date_of_birth")
        )
        passport_number = clean_text(
            get_field_value(ws, merged_lookup, excel_row, resolved_columns, "passport_number")
        )
        confirmation_number = clean_text(
            get_field_value(ws, merged_lookup, excel_row, resolved_columns, "confirmation_number")
        )

        rows.append(
            {
                "excel_row_number": excel_row,
                "source_row_number": source_row_number,
                "row_number_merge_anchor": row_number_merge_anchor,
                "group_label": group_label,
                "event_name": event_name,
                "first_name": first_name,
                "last_name": last_name,
                "full_name": " ".join([p for p in [first_name, last_name] if p]) or None,
                "mail": normalize_email(
                    get_field_value(ws, merged_lookup, excel_row, resolved_columns, "mail")
                ),
                "phone": normalize_phone(
                    get_field_value(ws, merged_lookup, excel_row, resolved_columns, "phone")
                ),
                "nationality": clean_text(
                    get_field_value(ws, merged_lookup, excel_row, resolved_columns, "nationality")
                ),
                "date_of_birth": date_of_birth,
                "passport_number": passport_number,
                "passport_expiration": normalize_date(
                    get_field_value(ws, merged_lookup, excel_row, resolved_columns, "passport_expiration")
                ),
                "remarks": clean_text(
                    get_field_value(ws, merged_lookup, excel_row, resolved_columns, "remarks")
                ),
                "meals": clean_text(
                    get_field_value(ws, merged_lookup, excel_row, resolved_columns, "meals")
                ),
                "food_restrictions": clean_text(
                    get_field_value(ws, merged_lookup, excel_row, resolved_columns, "food_restrictions")
                ),
                "confirmation_number": confirmation_number,
                "qty": qty,
                "qty_merge_anchor": qty_merge_anchor,
                "destination": destination,
                "hotel_name": hotel_name,
                "room": room,
                "check_in": check_in,
                "check_out": check_out,
                "nights": normalize_int(
                    get_field_value(ws, merged_lookup, excel_row, resolved_columns, "nights")
                ),
                "flights": extract_flight_segments_for_row(
                    ws=ws,
                    merged_lookup=merged_lookup,
                    excel_row=excel_row,
                    flight_column_map=flight_column_map,
                ),
                "passenger_key": build_passenger_key(
                    passport_number=passport_number,
                    last_name=last_name,
                    first_name=first_name,
                    date_of_birth=date_of_birth,
                    excel_row_number=excel_row,
                ),
                "profile_key": profile["key"],
            }
        )

    return rows