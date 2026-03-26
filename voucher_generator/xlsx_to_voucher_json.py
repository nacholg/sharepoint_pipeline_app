from __future__ import annotations

import argparse
import json
import re
import unicodedata
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook


# Operational spreadsheet layout
HEADER_ROW = 3
START_ROW = 4

# Internal field names -> accepted header aliases in Excel
FIELD_ALIASES: Dict[str, List[str]] = {
    "row_number": ["#", "NRO", "ROW NUMBER", "ROW #", "PASSENGER #"],
    "group_label": ["GROUP LABEL", "GROUP", "LABEL"],
    "last_name": ["TRAVELER LAST NAME", "LAST NAME", "SURNAME", "APELLIDO"],
    "first_name": ["TRAVELER FIRST NAME", "FIRST NAME", "NAME", "NOMBRE"],
    "mail": ["MAIL", "EMAIL", "E-MAIL"],
    "phone": ["TELEFONO", "TELEFONO ", "PHONE", "PHONE NUMBER", "CELLPHONE"],
    "nationality": ["NATIONALITY", "NACIONALIDAD"],
    "date_of_birth": ["DATE OF BIRTH", "DOB", "BIRTH DATE", "FECHA DE NACIMIENTO"],
    "passport_number": ["PASSPORT NUMBER", "PASSPORT", "PASSPORT NO", "NRO PASAPORTE"],
    "passport_expiration": [
        "EXPIRATION DATE",
        "EXPIRATION DATE ",
        "PASSPORT EXPIRATION",
        "PASSPORT EXPIRY",
        "VENCIMIENTO PASAPORTE",
    ],
    "remarks": ["REMARKS", "COMMENTS", "OBSERVATIONS", "OBSERVACIONES"],
    "food_restrictions": ["FOOD RESTRICTIONS", "FOOD", "DIETARY RESTRICTIONS"],
    "qty": ["QTY", "PAX", "PAX QTY", "QUANTITY"],
    "room": ["HAB", "ROOM", "ROOM CATEGORY", "ROOM TYPE"],
    "hotel_name": ["HOTEL NAME", "HOTEL", "HOTEL NAME RAW"],
    "destination": ["DESTINATION", "CITY", "DESTINO"],
    "check_in": ["CHECK IN HOTEL", "CHECK IN", "IN", "ARRIVAL"],
    "check_out": ["CHECK OUT HOTEL", "CHECK OUT", "OUT", "DEPARTURE"],
    "nights": ["ROOM NIGHTS", "NIGHTS", "NIGHTS QTY"],
}

# Required for this stage so the import fails early with a clear message
REQUIRED_FIELDS = [
    "row_number",
    "qty",
    "room",
    "hotel_name",
    "destination",
    "check_in",
    "check_out",
]


# -------- Normalization helpers --------
def clean_text(value: Any) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    if text in {"", "-", "--", "N/A", "n/a"}:
        return None
    return text


def normalize_email(value: Any) -> Optional[str]:
    text = clean_text(value)
    return text.lower() if text else None


def normalize_phone(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, (int, float)) and int(value) == value:
        return str(int(value))
    return clean_text(value)


def normalize_int(value: Any) -> Optional[int]:
    if value in (None, ""):
        return None
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return None


def normalize_date(value: Any) -> Optional[str]:
    if value in (None, ""):
        return None

    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")

    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")

    if isinstance(value, (int, float)):
        try:
            serial = float(value)
            if 1 <= serial <= 60000:
                excel_epoch = datetime(1899, 12, 30)
                parsed = excel_epoch + timedelta(days=serial)
                return parsed.strftime("%Y-%m-%d")
        except Exception:
            pass

    text = str(value).strip()
    for fmt in (
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%m/%d/%Y",
        "%d-%m-%Y",
        "%d.%m.%Y",
        "%Y/%m/%d",
    ):
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue

    return text


def normalize_key_text(value: Any) -> Optional[str]:
    text = clean_text(value)
    if not text:
        return None
    return text.upper()


def normalize_header(value: Any) -> Optional[str]:
    text = clean_text(value)
    if not text:
        return None
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.upper()
    text = re.sub(r"[^\w\s#]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text or None


def to_title_case(value: Any) -> Optional[str]:
    text = clean_text(value)
    if not text:
        return None
    return text.title()


def build_passenger_key(
    passport_number: Optional[str],
    last_name: Optional[str],
    first_name: Optional[str],
    date_of_birth: Optional[str],
    excel_row_number: int,
) -> str:
    passport = normalize_key_text(passport_number)
    if passport:
        return passport

    natural_key = "|".join(
        [
            normalize_key_text(last_name) or "NO-LASTNAME",
            normalize_key_text(first_name) or "NO-FIRSTNAME",
            normalize_date(date_of_birth) or "NO-DOB",
        ]
    )
    return f"{natural_key}|ROW-{excel_row_number}"


# -------- Excel merged-cell helpers --------
def build_merged_lookup(ws) -> Dict[Tuple[int, int], Tuple[int, int, int, int]]:
    lookup: Dict[Tuple[int, int], Tuple[int, int, int, int]] = {}
    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                lookup[(row, col)] = (min_row, min_col, max_row, max_col)
    return lookup


def get_effective_cell_value(ws, merged_lookup, row: int, col: int) -> Any:
    merge_bounds = merged_lookup.get((row, col))
    if merge_bounds:
        min_row, min_col, _, _ = merge_bounds
        return ws.cell(min_row, min_col).value
    return ws.cell(row, col).value


def get_merge_anchor_id(merged_lookup, row: int, col: Optional[int]) -> Optional[str]:
    if not col:
        return None
    merge_bounds = merged_lookup.get((row, col))
    if not merge_bounds:
        return None
    min_row, min_col, max_row, max_col = merge_bounds
    return f"R{min_row}C{min_col}:R{max_row}C{max_col}"


# -------- Header resolution --------
def build_header_index(
    ws,
    merged_lookup,
    header_row: int = HEADER_ROW,
) -> Dict[str, int]:
    header_index: Dict[str, int] = {}

    for col in range(1, ws.max_column + 1):
        raw_header = get_effective_cell_value(ws, merged_lookup, header_row, col)
        normalized = normalize_header(raw_header)
        if normalized and normalized not in header_index:
            header_index[normalized] = col

    return header_index


def resolve_columns(
    header_index: Dict[str, int],
    aliases: Dict[str, List[str]],
    required_fields: List[str],
) -> Dict[str, int]:
    resolved: Dict[str, int] = {}
    missing_required: List[str] = []

    for field_name, candidates in aliases.items():
        for candidate in candidates:
            normalized_candidate = normalize_header(candidate)
            if normalized_candidate and normalized_candidate in header_index:
                resolved[field_name] = header_index[normalized_candidate]
                break

        if field_name in required_fields and field_name not in resolved:
            missing_required.append(field_name)

    if missing_required:
        details = []
        for field_name in missing_required:
            aliases_text = ", ".join(aliases.get(field_name, []))
            details.append(f"- {field_name}: expected one of [{aliases_text}]")
        raise ValueError(
            "Missing required columns in header row "
            f"{HEADER_ROW}:\n" + "\n".join(details)
        )

    return resolved


def get_field_value(
    ws,
    merged_lookup,
    row: int,
    resolved_columns: Dict[str, int],
    field_name: str,
) -> Any:
    col = resolved_columns.get(field_name)
    if not col:
        return None
    return get_effective_cell_value(ws, merged_lookup, row, col)


# -------- Row extraction --------
def read_effective_rows(
    xlsx_path: Path,
    sheet_name: Optional[str] = None,
    start_row: int = START_ROW,
    header_row: int = HEADER_ROW,
) -> List[Dict[str, Any]]:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    merged_lookup = build_merged_lookup(ws)

    header_index = build_header_index(ws, merged_lookup, header_row=header_row)
    resolved_columns = resolve_columns(
        header_index=header_index,
        aliases=FIELD_ALIASES,
        required_fields=REQUIRED_FIELDS,
    )

    rows: List[Dict[str, Any]] = []

    for excel_row in range(start_row, ws.max_row + 1):
        raw_first_name = get_field_value(ws, merged_lookup, excel_row, resolved_columns, "first_name")
        raw_last_name = get_field_value(ws, merged_lookup, excel_row, resolved_columns, "last_name")
        raw_destination = get_field_value(ws, merged_lookup, excel_row, resolved_columns, "destination")
        raw_hotel_name = get_field_value(ws, merged_lookup, excel_row, resolved_columns, "hotel_name")
        raw_room = get_field_value(ws, merged_lookup, excel_row, resolved_columns, "room")
        raw_check_in = get_field_value(ws, merged_lookup, excel_row, resolved_columns, "check_in")
        raw_check_out = get_field_value(ws, merged_lookup, excel_row, resolved_columns, "check_out")
        raw_qty = get_field_value(ws, merged_lookup, excel_row, resolved_columns, "qty")

        first_name = clean_text(raw_first_name)
        last_name = clean_text(raw_last_name)
        destination = clean_text(raw_destination)
        hotel_name = clean_text(raw_hotel_name)
        room = clean_text(raw_room)
        check_in = normalize_date(raw_check_in)
        check_out = normalize_date(raw_check_out)
        qty = normalize_int(raw_qty)

        has_passenger = bool(first_name or last_name)
        has_voucher_context = any([destination, hotel_name, room, check_in, check_out, qty])

        if not has_passenger and not has_voucher_context:
            continue

        qty_merge_anchor = get_merge_anchor_id(
            merged_lookup,
            excel_row,
            resolved_columns.get("qty"),
        )
        row_number_merge_anchor = get_merge_anchor_id(
            merged_lookup,
            excel_row,
            resolved_columns.get("row_number"),
        )

        source_row_number = normalize_int(
            get_field_value(ws, merged_lookup, excel_row, resolved_columns, "row_number")
        )
        group_label = clean_text(
            get_field_value(ws, merged_lookup, excel_row, resolved_columns, "group_label")
        )
        date_of_birth = normalize_date(
            get_field_value(ws, merged_lookup, excel_row, resolved_columns, "date_of_birth")
        )
        passport_number = clean_text(
            get_field_value(ws, merged_lookup, excel_row, resolved_columns, "passport_number")
        )

        rows.append(
            {
                "excel_row_number": excel_row,
                "source_row_number": source_row_number,
                "row_number_merge_anchor": row_number_merge_anchor,
                "group_label": group_label,
                "first_name": first_name,
                "last_name": last_name,
                "full_name": " ".join([p for p in [first_name, last_name] if p]) or None,
                "mail": normalize_email(
                    get_field_value(ws, merged_lookup, excel_row, resolved_columns, "mail")
                ),
                "phone": normalize_phone(
                    get_field_value(ws, merged_lookup, excel_row, resolved_columns, "phone")
                ),
                "nationality": clean_text(
                    get_field_value(ws, merged_lookup, excel_row, resolved_columns, "nationality")
                ),
                "date_of_birth": date_of_birth,
                "passport_number": passport_number,
                "passport_expiration": normalize_date(
                    get_field_value(ws, merged_lookup, excel_row, resolved_columns, "passport_expiration")
                ),
                "remarks": clean_text(
                    get_field_value(ws, merged_lookup, excel_row, resolved_columns, "remarks")
                ),
                "food_restrictions": clean_text(
                    get_field_value(ws, merged_lookup, excel_row, resolved_columns, "food_restrictions")
                ),
                "qty": qty,
                "qty_merge_anchor": qty_merge_anchor,
                "destination": destination,
                "hotel_name": hotel_name,
                "room": room,
                "check_in": check_in,
                "check_out": check_out,
                "nights": normalize_int(
                    get_field_value(ws, merged_lookup, excel_row, resolved_columns, "nights")
                ),
                "passenger_key": build_passenger_key(
                    passport_number=passport_number,
                    last_name=last_name,
                    first_name=first_name,
                    date_of_birth=date_of_birth,
                    excel_row_number=excel_row,
                ),
            }
        )

    return rows


# -------- Logical voucher grouping --------
def build_voucher_blocks(rows: List[Dict[str, Any]]) -> List[List[Dict[str, Any]]]:
    blocks: List[List[Dict[str, Any]]] = []
    current_block: List[Dict[str, Any]] = []
    current_block_key: Optional[str] = None

    for row in rows:
        qty = row.get("qty")
        qty_anchor = row.get("qty_merge_anchor")

        if qty_anchor:
            block_key = f"MERGED:{qty_anchor}"
            if current_block and current_block_key != block_key:
                blocks.append(current_block)
                current_block = []
            current_block_key = block_key
            current_block.append(row)
            continue

        if current_block:
            blocks.append(current_block)
            current_block = []
            current_block_key = None

        if qty is not None or row.get("full_name") or row.get("hotel_name") or row.get("destination"):
            blocks.append([row])

    if current_block:
        blocks.append(current_block)

    return blocks


def dedupe_real_passengers(rows: Iterable[Dict[str, Any]]) -> List[Dict[str, Any]]:
    seen: set[str] = set()
    passengers: List[Dict[str, Any]] = []

    for row in sorted(rows, key=lambda r: (r["excel_row_number"], r.get("source_row_number") or 0)):
        if not row.get("full_name"):
            continue

        passenger_key = row["passenger_key"]
        if passenger_key in seen:
            continue
        seen.add(passenger_key)

        passengers.append(
            {
                "passenger_key": passenger_key,
                "first_name": row["first_name"],
                "last_name": row["last_name"],
                "full_name": row["full_name"],
                "mail": row["mail"],
                "phone": row["phone"],
                "nationality": row["nationality"],
                "date_of_birth": row["date_of_birth"],
                "passport_number": row["passport_number"],
                "passport_expiration": row["passport_expiration"],
                "food_restrictions": row["food_restrictions"],
                "remarks": row["remarks"],
            }
        )

    return passengers


def pad_passengers(passengers: List[Dict[str, Any]], qty: int, block_rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    padded = list(passengers)
    next_index = len(padded) + 1
    while len(padded) < qty:
        padded.append(
            {
                "passenger_key": f"PENDING-{block_rows[0]['excel_row_number']}-{next_index}",
                "first_name": None,
                "last_name": None,
                "full_name": "NAME PENDING",
                "mail": None,
                "phone": None,
                "nationality": None,
                "date_of_birth": None,
                "passport_number": None,
                "passport_expiration": None,
                "food_restrictions": None,
                "remarks": None,
            }
        )
        next_index += 1
    return padded


def choose_block_header(block_rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    return sorted(block_rows, key=lambda r: r["excel_row_number"])[0]


def build_voucher_payloads(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    blocks = build_voucher_blocks(rows)
    payloads: List[Dict[str, Any]] = []

    for voucher_index, block_rows in enumerate(blocks, start=1):
        block_rows = sorted(block_rows, key=lambda r: r["excel_row_number"])
        header = choose_block_header(block_rows)
        passengers = dedupe_real_passengers(block_rows)

        declared_qty = header.get("qty") or 0
        is_merged_qty_block = bool(header.get("qty_merge_anchor"))

        if is_merged_qty_block:
            pax_count = max(1, len(block_rows))
        else:
            pax_count = declared_qty if declared_qty > 0 else max(1, len(passengers))

        passengers = pad_passengers(passengers, pax_count, block_rows)

        voucher_id = header.get("source_row_number") or voucher_index

        payloads.append(
            {
                "voucher_id": voucher_id,
                "voucher_group_key": f"VOUCHER-{voucher_id}",
                "voucher": {
                    "voucher_code": None,
                    "confirmation_number": None,
                    "issue_date": None,
                    "remarks": None,
                    "meals": None,
                    "other_services": None,
                },
                "source": {
                    "sheet_name": None,
                    "group_label": header.get("group_label"),
                    "excel_rows": [r["excel_row_number"] for r in block_rows],
                    "row_numbers": [
                        r["source_row_number"] for r in block_rows if r.get("source_row_number") is not None
                    ],
                    "qty_merge_anchor": header.get("qty_merge_anchor"),
                },
                "destination": {
                    "name": header.get("destination"),
                    "display_name": to_title_case(header.get("destination")) or header.get("destination"),
                },
                "hotel": {
                    "name": header.get("hotel_name"),
                    "display_name": to_title_case(header.get("hotel_name")) or header.get("hotel_name"),
                    "address": None,
                    "city": None,
                    "country": None,
                    "phone": None,
                },
                "stay": {
                    "check_in": header.get("check_in"),
                    "check_out": header.get("check_out"),
                    "nights": header.get("nights"),
                },
                "rooms": [
                    {
                        "room_sequence": 1,
                        "room_count": 1,
                        "room_category": header.get("room"),
                        "additional_info": None,
                        "pax_count": pax_count,
                    }
                ],
                "passengers": passengers,
            }
        )

    payloads.sort(
        key=lambda p: (
            str(p.get("voucher_id") or ""),
            p["stay"]["check_in"] or "",
            p["destination"]["name"] or "",
            p["hotel"]["name"] or "",
        )
    )
    return payloads


# -------- CLI --------
def main() -> None:
    parser = argparse.ArgumentParser(
        description="Convert the operational XLSX into voucher JSON payloads, honoring merged QTY blocks."
    )
    parser.add_argument("input", help="Path to the source .xlsx file")
    parser.add_argument("-o", "--output", help="Path to the output .json file")
    parser.add_argument("--sheet", help="Optional sheet name. Defaults to the first sheet.")
    parser.add_argument("--pretty", action="store_true", help="Pretty-print the JSON output.")
    parser.add_argument(
        "--debug-rows",
        action="store_true",
        help="Also emit a .rows.json file with normalized rows for debugging.",
    )
    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output) if args.output else input_path.with_suffix(".voucher_payloads.json")

    rows = read_effective_rows(input_path, sheet_name=args.sheet)
    payloads = build_voucher_payloads(rows)

    output_path.write_text(
        json.dumps(payloads, ensure_ascii=False, indent=2 if args.pretty else None),
        encoding="utf-8",
    )

    if args.debug_rows:
        debug_path = output_path.with_suffix(".rows.json")
        debug_path.write_text(
            json.dumps(rows, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        print(f"Debug rows: {debug_path}")

    print(f"Rows processed: {len(rows)}")
    print(f"Voucher payloads generated: {len(payloads)}")
    print(f"Output: {output_path}")


if __name__ == "__main__":
    main()